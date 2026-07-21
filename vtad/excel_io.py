import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .censys_client import format_services
from .utils import is_ip

DOMAIN_PATTERN = re.compile(
    r"^[a-zA-Z0-9](?:[a-zA-Z0-9-]*[a-zA-Z0-9])?(?:\.[a-zA-Z0-9](?:[a-zA-Z0-9-]*[a-zA-Z0-9])?)+$"
)

HEADER = [
    "Hedef",
    "Tür",
    "Çözümlenen IP",
    "VT Bulundu",
    "VT Zararlı",
    "VT Şüpheli",
    "VT Zararsız",
    "VT Tespitsiz",
    "VT Zararlı Motorlar",
    "AbuseIPDB Skoru",
    "AbuseIPDB Rapor Sayısı",
    "AbuseIPDB Kategoriler",
    "AbuseIPDB Kullanım Tipi",
    "AbuseIPDB Hostname",
    "RDAP Kayıt Tarihi",
    "RDAP Domain Yaşı (gün)",
    "RDAP Registrar",
    "RDAP Kayıt Sahibi",
    "RDAP Nameserver",
    "Censys Açık Servisler",
    "Ülke",
    "Şehir",
    "ASN / Operatör",
    "Koordinatlar",
    "Sonuç",
    "Not",
]

_VERDICT_STYLES = {
    "bad": (PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006")),
    "warn": (PatternFill("solid", fgColor="FFEB9C"), Font(color="9C6500")),
    "ok": (PatternFill("solid", fgColor="C6EFCE"), Font(color="006100")),
}


def _looks_like_target(value: str) -> bool:
    return is_ip(value) or bool(DOMAIN_PATTERN.match(value))


def read_targets_from_excel(path: str) -> list[str]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active

    targets = []
    is_first_value = True
    for (value,) in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
        if value is None:
            continue
        value = str(value).strip()
        if not value:
            continue
        if is_first_value:
            is_first_value = False
            if not _looks_like_target(value):
                continue  # başlık satırı
        targets.append(value)

    workbook.close()
    return targets


def write_results_to_excel(path: str, results: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sonuçlar"

    sheet.append(HEADER)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    verdict_col = HEADER.index("Sonuç") + 1

    for result in results:
        vt_result = result.get("vt_result") or {}
        stats = vt_result.get("stats", {}) if vt_result.get("found") else {}
        abuse_result = result.get("abuse_result")
        rdap_result = result.get("rdap_result")
        censys_result = result.get("censys_result")

        notes = []
        if result.get("vt_error"):
            notes.append(f"VT: {result['vt_error']}")
        if result.get("abuse_error"):
            notes.append(f"AbuseIPDB: {result['abuse_error']}")
        if rdap_result and rdap_result.get("error"):
            notes.append(f"RDAP: {rdap_result['error']}")
        if result.get("censys_error"):
            notes.append(f"Censys: {result['censys_error']}")

        censys_services = ""
        if censys_result and censys_result.get("found"):
            censys_services = format_services(censys_result.get("services", []))

        top_categories = ""
        if abuse_result and abuse_result.get("top_categories"):
            top_categories = ", ".join(
                f"{name} ({count})" for name, count in abuse_result["top_categories"]
            )

        # Konum ve ağ bilgisini kaynaklardan birleştir
        censys_result = censys_result or {}
        abuse_data = abuse_result or {}
        rdap_data = rdap_result or {}
        country = censys_result.get("country") or abuse_data.get("country_name") or ""
        city = censys_result.get("city") or ""
        asn = censys_result.get("asn") or vt_result.get("asn")
        as_name = censys_result.get("as_name") or vt_result.get("as_owner") or abuse_data.get("isp") or ""
        asn_operator = f"AS{asn} {as_name}".strip() if asn else as_name

        sheet.append([
            result["target"],
            "IP" if result["is_ip"] else "Domain",
            result.get("resolved_ip") or "",
            "Evet" if vt_result.get("found") else "Hayır",
            stats.get("malicious", ""),
            stats.get("suspicious", ""),
            stats.get("harmless", ""),
            stats.get("undetected", ""),
            ", ".join(vt_result.get("malicious_engines", [])),
            abuse_data.get("abuse_score", "") if abuse_result else "",
            abuse_data.get("total_reports", "") if abuse_result else "",
            top_categories,
            abuse_data.get("usage_type") or "",
            ", ".join(abuse_data.get("hostnames", [])),
            rdap_data.get("creation_date") or "",
            rdap_data.get("age_days") if rdap_data.get("age_days") is not None else "",
            rdap_data.get("registrar") or "",
            rdap_data.get("registrant_org") or rdap_data.get("registrant_name") or "",
            ", ".join(rdap_data.get("nameservers", [])),
            censys_services,
            country,
            city,
            asn_operator,
            censys_result.get("coordinates") or "",
            result["verdict_text"],
            "; ".join(notes),
        ])

        fill, font = _VERDICT_STYLES[result["verdict_kind"]]
        verdict_cell = sheet.cell(row=sheet.max_row, column=verdict_col)
        verdict_cell.fill = fill
        verdict_cell.font = font

    for idx, header in enumerate(HEADER, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = max(14, len(header) + 4)

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"

    workbook.save(path)
